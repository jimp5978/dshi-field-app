#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import sys
import json
import os
from openpyxl import load_workbook

def parse_excel_file(file_path):
    """Excel 파일에서 A열의 Assembly Code들을 추출"""
    try:
        if not os.path.exists(file_path):
            return {
                'success': False,
                'error': f'파일을 찾을 수 없습니다: {file_path}'
            }
        
        # Excel 파일 로드
        workbook = load_workbook(file_path)
        worksheet = workbook.active  # 첫 번째 시트
        
        # A열에서 Assembly Code 추출 (A1부터)
        assembly_codes = []
        for row in worksheet.iter_rows(min_col=1, max_col=1, values_only=True):
            cell_value = row[0]
            if cell_value and str(cell_value).strip():
                assembly_codes.append(str(cell_value).strip())
        
        # 100개 제한 확인
        if len(assembly_codes) > 100:
            return {
                'success': False,
                'error': f'최대 100개까지만 업로드 가능합니다. 현재: {len(assembly_codes)}개'
            }
        
        if len(assembly_codes) == 0:
            return {
                'success': False,
                'error': 'A열에서 Assembly Code를 찾을 수 없습니다'
            }
        
        return {
            'success': True,
            'assembly_codes': assembly_codes,
            'count': len(assembly_codes)
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': f'Excel 파일 파싱 오류: {str(e)}'
        }

if __name__ == '__main__':
    if len(sys.argv) != 2:
        result = {
            'success': False,
            'error': '사용법: python parse_excel.py <excel_file_path>'
        }
    else:
        file_path = sys.argv[1]
        result = parse_excel_file(file_path)
    
    # JSON 출력 (Ruby에서 읽을 수 있도록)
    print(json.dumps(result, ensure_ascii=False))