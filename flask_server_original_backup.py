from flask import Flask, request, jsonify
from flask_cors import CORS
import mysql.connector
from mysql.connector import Error
import hashlib
import jwt
import datetime
from functools import wraps
import os
import json
import logging
from datetime import datetime as dt
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import decimal

app = Flask(__name__)
CORS(app)

# JSON ì§ë ¬í™”ë¥¼ ìœ„í•œ ì»¤ìŠ¤í…€ ì¸ì½”ë”
class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime.datetime, datetime.date)):
            return obj.isoformat()
        elif isinstance(obj, decimal.Decimal):
            return float(obj)
        elif isinstance(obj, bytes):
            return obj.decode('utf-8')
        return super().default(obj)

app.json_encoder = CustomJSONEncoder

# JWT ì„¤ì •
app.config['SECRET_KEY'] = 'dshi-field-pad-secret-key-2025'

# ë””ë²„ê·¸ ë¡œê¹… ì„¤ì •
logging.basicConfig(
    level=logging.DEBUG,
    format='ğŸ› DEBUG [%(asctime)s]: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.FileHandler('flask_debug.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# MySQL ì—°ê²° ì„¤ì •
from config_env import get_db_config, get_server_config
DB_CONFIG = get_db_config()
SERVER_CONFIG = get_server_config()

def get_db_connection():
    """ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° í•¨ìˆ˜"""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        return connection
    except Error as e:
        print(f"ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì˜¤ë¥˜: {e}")
        return None

def calculate_assembly_status(assembly_data):
    """ì¡°ë¦½í’ˆ ìƒíƒœ ê³„ì‚° (Ruby ProcessManager ë¡œì§ê³¼ ë™ì¼)"""
    try:
        # 8ë‹¨ê³„ ê³µì • ìˆœì„œ
        processes = [
            ('FIT_UP', assembly_data.get('fit_up_date')),
            ('FINAL', assembly_data.get('final_date')),
            ('ARUP_FINAL', assembly_data.get('arup_final_date')),
            ('GALV', assembly_data.get('galv_date')),
            ('ARUP_GALV', assembly_data.get('arup_galv_date')),
            ('SHOT', assembly_data.get('shot_date')),
            ('PAINT', assembly_data.get('paint_date')),
            ('ARUP_PAINT', assembly_data.get('arup_paint_date'))
        ]
        
        # ì™„ë£Œëœ ê³µì •ë“¤ê³¼ ë¶ˆí•„ìš”í•œ ê³µì •ë“¤ êµ¬ë¶„
        completed_processes = []
        skipped_processes = []
        
        for name, date in processes:
            if date and str(date).strip():
                date_str = str(date)
                if '1900' in date_str:
                    # 1900-01-01ì€ ë¶ˆí•„ìš”í•œ ê³µì • (ê±´ë„ˆë›°ê¸°)
                    skipped_processes.append(name)
                else:
                    # ì‹¤ì œ ì™„ë£Œëœ ê³µì •
                    completed_processes.append((name, date))
        
        # ì „ì²´ ê³µì • ìˆ˜ (8ê°œ) - ê±´ë„ˆë›´ ê³µì • ìˆ˜ = í•„ìš”í•œ ê³µì • ìˆ˜
        total_required_processes = 8 - len(skipped_processes)
        
        # ìƒíƒœ ë° ë§ˆì§€ë§‰ ê³µì • ê³„ì‚°
        if completed_processes:
            # ê°€ì¥ ë§ˆì§€ë§‰ ì™„ë£Œëœ ê³µì •
            last_process_name, last_date = completed_processes[-1]
            
            # ì‹¤ì œ ì™„ë£Œëœ ê³µì • ìˆ˜ê°€ í•„ìš”í•œ ê³µì • ìˆ˜ì™€ ê°™ìœ¼ë©´ ì™„ë£Œ
            status = 'ì™„ë£Œ' if len(completed_processes) >= total_required_processes else 'ì§„í–‰ì¤‘'
            last_process = last_process_name
        else:
            last_process = 'ì‹œì‘ì „'
            status = 'ëŒ€ê¸°'
        
        # ë‹¤ìŒ ê³µì • ê³„ì‚°
        next_process = None
        for name, date in processes:
            if date and str(date).strip():
                date_str = str(date)
                if '1900' in date_str:
                    # ë¶ˆí•„ìš”í•œ ê³µì •ì€ ê±´ë„ˆë›°ê¸°
                    continue
            else:
                # ë‚ ì§œê°€ ì—†ê±°ë‚˜ ë¹„ì–´ìˆëŠ” ê²½ìš° ë¯¸ì™„ë£Œ ê³µì •
                next_process = name
                break
        
        # ë‹¤ìŒ ê³µì • í•œêµ­ì–´ ë³€í™˜
        next_process_korean = {
            'FIT_UP': 'FIT-UP',
            'FINAL': 'FINAL',
            'ARUP_FINAL': 'ARUP FINAL',
            'GALV': 'GALV',
            'ARUP_GALV': 'ARUP GALV',
            'SHOT': 'SHOT',
            'PAINT': 'PAINT',
            'ARUP_PAINT': 'ARUP PAINT'
        }.get(next_process, 'ì™„ë£Œ')
        
        # ê³„ì‚°ëœ ìƒíƒœ ì •ë³´ë¥¼ assembly_dataì— ì¶”ê°€
        assembly_data['status'] = status
        assembly_data['lastProcess'] = last_process
        assembly_data['nextProcess'] = next_process_korean
        
        return assembly_data
        
    except Exception as e:
        print(f"ìƒíƒœ ê³„ì‚° ì˜¤ë¥˜: {e}")
        # ì˜¤ë¥˜ ì‹œ ê¸°ë³¸ê°’ ì„¤ì •
        assembly_data['status'] = 'ì˜¤ë¥˜'
        assembly_data['lastProcess'] = 'ì•Œ ìˆ˜ ì—†ìŒ'
        assembly_data['nextProcess'] = 'ì•Œ ìˆ˜ ì—†ìŒ'
        return assembly_data

def get_user_info(user_id):
    """ì‚¬ìš©ì ì •ë³´ ì¡°íšŒ í—¬í¼ í•¨ìˆ˜"""
    try:
        connection = get_db_connection()
        if not connection:
            return None
        
        cursor = connection.cursor(dictionary=True)
        try:
            cursor.execute("""
                SELECT id, username, full_name, permission_level, company
                FROM users 
                WHERE id = %s AND is_active = TRUE
            """, (user_id,))
        except:
            # company ì»¬ëŸ¼ì´ ì—†ëŠ” ê²½ìš°
            cursor.execute("""
                SELECT id, username, full_name, permission_level
                FROM users 
                WHERE id = %s AND is_active = TRUE
            """, (user_id,))
        
        user = cursor.fetchone()
        if user and 'company' not in user:
            user['company'] = ''
        cursor.close()
        connection.close()
        return user
    except Exception as e:
        print(f"ì‚¬ìš©ì ì •ë³´ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return None

def token_required(f):
    """JWT í† í° ê²€ì¦ ë°ì½”ë ˆì´í„°"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        
        if not token:
            return jsonify({'message': 'í† í°ì´ ì—†ìŠµë‹ˆë‹¤'}), 401
        
        try:
            if token.startswith('Bearer '):
                token = token[7:]
            
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            # Excel ì—…ë¡œë“œ APIëŠ” ì „ì²´ JWT ë°ì´í„°ê°€ í•„ìš”í•˜ë¯€ë¡œ í•¨ìˆ˜ë³„ë¡œ êµ¬ë¶„
            if f.__name__ in ['upload_excel', 'upload_assembly_codes']:
                current_user = data  # ì „ì²´ JWT ë°ì´í„° ì „ë‹¬
            else:
                current_user = data['user_id']  # ê¸°ì¡´ ë°©ì‹ ìœ ì§€
        except:
            return jsonify({'message': 'í† í°ì´ ìœ íš¨í•˜ì§€ ì•ŠìŠµë‹ˆë‹¤'}), 401
        
        return f(current_user, *args, **kwargs)
    return decorated

@app.route('/api/login', methods=['POST'])
def login():
    """ì‚¬ìš©ì ë¡œê·¸ì¸ - ë°ì´í„°ë² ì´ìŠ¤ ê¸°ë°˜ (í•˜ë“œì½”ë”© ë°±ì—…)"""
    try:
        data = request.get_json()
        username = data.get('username')
        password_hash = data.get('password_hash')
        
        if not username or not password_hash:
            return jsonify({'success': False, 'message': 'ì•„ì´ë””ì™€ ë¹„ë°€ë²ˆí˜¸ë¥¼ ì…ë ¥í•˜ì„¸ìš”'}), 400
        
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        try:
            cursor = connection.cursor(dictionary=True)
            
            # ë°ì´í„°ë² ì´ìŠ¤ì—ì„œ ì‚¬ìš©ì ì¡°íšŒ
            try:
                cursor.execute("""
                    SELECT id, username, password_hash, full_name, permission_level, company, is_active
                    FROM users 
                    WHERE username = %s AND is_active = TRUE
                """, (username,))
            except:
                # company ì»¬ëŸ¼ì´ ì—†ëŠ” ê²½ìš°
                cursor.execute("""
                    SELECT id, username, password_hash, full_name, permission_level, is_active
                    FROM users 
                    WHERE username = %s AND is_active = TRUE
                """, (username,))
            
            user = cursor.fetchone()
            # company ì»¬ëŸ¼ì´ ì—†ëŠ” ê²½ìš° ê¸°ë³¸ê°’ ì„¤ì •
            if user and 'company' not in user:
                user['company'] = ''
            cursor.close()
            connection.close()
            
        except Exception as e:
            print(f"ë°ì´í„°ë² ì´ìŠ¤ ì¡°íšŒ ì˜¤ë¥˜: {e}")
            connection.close()
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì˜¤ë¥˜'}), 500
        
        # ì‚¬ìš©ì ì¸ì¦
        if user and user['password_hash'] == password_hash:
            # JWT í† í° ìƒì„±
            token = jwt.encode({
                'user_id': user['id'],
                'username': user['username'],
                'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=24)
            }, app.config['SECRET_KEY'], algorithm='HS256')
            
            return jsonify({
                'success': True,
                'message': 'ë¡œê·¸ì¸ ì„±ê³µ',
                'token': token,
                'user': {
                    'id': user['id'],
                    'username': user['username'],
                    'full_name': user['full_name'],
                    'permission_level': user['permission_level'],
                    'company': user.get('company', '')
                }
            })
        else:
            return jsonify({'success': False, 'message': 'ì•„ì´ë”” ë˜ëŠ” ë¹„ë°€ë²ˆí˜¸ê°€ í‹€ë ¸ìŠµë‹ˆë‹¤'}), 401
            
    except Exception as e:
        print(f"ë¡œê·¸ì¸ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/assemblies', methods=['GET'])
def get_assemblies():
    """ì¡°ë¦½í’ˆ ëª©ë¡ ì¡°íšŒ"""
    try:
        search = request.args.get('search', '')
        
        if not search:
            return jsonify({'success': False, 'message': 'ê²€ìƒ‰ì–´ë¥¼ ì…ë ¥í•˜ì„¸ìš”'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # assembly_code ë 3ìë¦¬ ìˆ«ìë¡œ ê²€ìƒ‰
        cursor.execute("""
            SELECT 
                id,
                assembly_code,
                zone,
                item,
                fit_up_date,
                nde_date,
                vidi_date,
                galv_date,
                shot_date,
                paint_date,
                packing_date
            FROM assembly_items 
            WHERE RIGHT(assembly_code, 3) = %s
            ORDER BY assembly_code
            LIMIT 50
        """, (search,))
        
        rows = cursor.fetchall()
        cursor.close()
        connection.close()
        
        # ë°ì´í„° ë³€í™˜ (ì•± í˜•ì‹ì— ë§ê²Œ)
        assemblies = []
        for row in rows:
            # ìµœì¢… ì™„ë£Œëœ ê³µì • ì°¾ê¸°
            processes = [
                ('Fit-up', row['fit_up_date']),
                ('NDE', row['nde_date']),
                ('VIDI', row['vidi_date']),
                ('GALV', row['galv_date']),
                ('SHOT', row['shot_date']),
                ('PAINT', row['paint_date']),
                ('PACKING', row['packing_date'])
            ]
            
            # ì™„ë£Œëœ ê³µì •ë“¤ë§Œ í•„í„°ë§ (Noneê³¼ 1900-01-01 ì œì™¸)
            completed_processes = []
            for name, date in processes:
                if date is not None and date != datetime.date(1900, 1, 1):
                    completed_processes.append((name, date))
            
            if completed_processes:
                # ê°€ì¥ ë§ˆì§€ë§‰ ì™„ë£Œëœ ê³µì •
                last_process_name, last_date = completed_processes[-1]
                status = 'ì™„ë£Œ' if len(completed_processes) == 7 else 'ì§„í–‰ì¤‘'
                completed_date = last_date.strftime('%Y-%m-%d') if last_date else ''
            else:
                last_process_name = 'ì‹œì‘ì „'
                status = 'ëŒ€ê¸°'
                completed_date = ''
            
            assemblies.append({
                'id': str(row['id']),
                'name': row['assembly_code'],
                'location': row['zone'] or '',
                'status': status,
                'drawing_number': row['item'] or '',
                'lastProcess': last_process_name,
                'completedDate': completed_date
            })
        
        return jsonify({
            'success': True,
            'assemblies': assemblies
        })
        
    except Exception as e:
        print(f"ê²€ìƒ‰ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/assemblies/search', methods=['GET'])
def search_assemblies():
    """ì¡°ë¦½í’ˆ ê²€ìƒ‰"""
    try:
        query = request.args.get('q', '')
        
        if not query:
            return jsonify({'success': False, 'message': 'ê²€ìƒ‰ì–´ë¥¼ ì…ë ¥í•˜ì„¸ìš”'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        search_pattern = f"%{query}%"
        
        # ìˆ«ìì¸ì§€ í™•ì¸í•˜ì—¬ ë 3ìë¦¬ ê²€ìƒ‰ ë˜ëŠ” ì¼ë°˜ ê²€ìƒ‰ ì ìš©
        if query.isdigit() and len(query) <= 3:
            # ë 3ìë¦¬ ìˆ«ì ê²€ìƒ‰
            cursor.execute("""
                SELECT assembly_code, company, zone, item, weight_gross,
                       fit_up_date, final_date, arup_final_date, galv_date, 
                       arup_galv_date, shot_date, paint_date, arup_paint_date
                FROM arup_ecs 
                WHERE RIGHT(assembly_code, 3) = %s
                ORDER BY assembly_code
                LIMIT 50
            """, (query.zfill(3),))  # 3ìë¦¬ë¡œ íŒ¨ë”© (ì˜ˆ: "27" -> "027")
        else:
            # ì¼ë°˜ ê²€ìƒ‰ (assembly_codeë‚˜ itemì— í¬í•¨ëœ ê²½ìš°)
            search_pattern = f"%{query}%"
            cursor.execute("""
                SELECT assembly_code, company, zone, item, weight_gross,
                       fit_up_date, final_date, arup_final_date, galv_date, 
                       arup_galv_date, shot_date, paint_date, arup_paint_date
                FROM arup_ecs 
                WHERE assembly_code LIKE %s OR item LIKE %s
                ORDER BY assembly_code
                LIMIT 50
            """, (search_pattern, search_pattern))
        
        assemblies = cursor.fetchall()
        cursor.close()
        connection.close()
        
        # ê° ì¡°ë¦½í’ˆì— ëŒ€í•´ ìƒíƒœ ê³„ì‚° (ê³µí†µ í•¨ìˆ˜ ì‚¬ìš©)
        processed_assemblies = []
        for assembly in assemblies:
            # dictionaryë¥¼ calculate_assembly_statusì— ì „ë‹¬í•  ìˆ˜ ìˆë„ë¡ ë³€í™˜
            assembly_dict = dict(assembly)
            processed_assembly = calculate_assembly_status(assembly_dict)
            processed_assemblies.append(processed_assembly)
        
        return jsonify({
            'success': True,
            'data': processed_assemblies
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/inspection-requests', methods=['POST'])
@token_required
def create_inspection_request(current_user):
    """ê²€ì‚¬ì‹ ì²­ ìƒì„±"""
    try:
        data = request.get_json()
        assembly_codes = data.get('assembly_codes', [])
        inspection_type = data.get('inspection_type')
        request_date = data.get('request_date')
        
        if not assembly_codes or not inspection_type or not request_date:
            return jsonify({'success': False, 'message': 'í•„ìˆ˜ ë°ì´í„°ê°€ ëˆ„ë½ë˜ì—ˆìŠµë‹ˆë‹¤'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # ì‚¬ìš©ì ì •ë³´ ì¡°íšŒ
        user_info = get_user_info(current_user)
        if not user_info:
            return jsonify({'success': False, 'message': 'ì‚¬ìš©ì ì •ë³´ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        username = user_info['username']
        full_name = user_info['full_name']
        
        # ì¤‘ë³µ ì²´í¬ ë° ê²€ì‚¬ì‹ ì²­ ì €ì¥
        inserted_count = 0
        duplicate_items = []
        
        for assembly_code in assembly_codes:
            # ì„ ì°©ìˆœ ì²´í¬: ê°™ì€ ASSEMBLY + ê°™ì€ ê²€ì‚¬íƒ€ì…ì´ ì´ë¯¸ ìˆëŠ”ì§€ í™•ì¸ (ì·¨ì†Œëœ í•­ëª© ì œì™¸)
            cursor.execute("""
                SELECT id, requested_by_name, request_date 
                FROM inspection_requests 
                WHERE assembly_code = %s AND inspection_type = %s AND status != 'ì·¨ì†Œë¨'
                LIMIT 1
            """, (assembly_code, inspection_type))
            
            existing_request = cursor.fetchone()
            
            if existing_request:
                # ì´ë¯¸ ì‹ ì²­ëœ í•­ëª©
                duplicate_items.append({
                    'assembly_code': assembly_code,
                    'existing_requester': existing_request['requested_by_name'],
                    'existing_date': existing_request['request_date'].strftime('%Y-%m-%d')
                })
            else:
                # ìƒˆë¡œ ì‹ ì²­ ê°€ëŠ¥í•œ í•­ëª©
                cursor.execute("""
                    INSERT INTO inspection_requests (
                        assembly_code, 
                        inspection_type, 
                        requested_by_user_id, 
                        requested_by_name, 
                        request_date
                    ) VALUES (%s, %s, %s, %s, %s)
                """, (assembly_code, inspection_type, current_user, full_name, request_date))
                inserted_count += 1
        
        connection.commit()
        cursor.close()
        connection.close()
        
        # ê²°ê³¼ ë©”ì‹œì§€ ìƒì„±
        if inserted_count > 0 and len(duplicate_items) == 0:
            # ëª¨ë‘ ì„±ê³µ
            return jsonify({
                'success': True,
                'message': f'{inserted_count}ê°œ í•­ëª©ì˜ {inspection_type} ê²€ì‚¬ê°€ ì‹ ì²­ë˜ì—ˆìŠµë‹ˆë‹¤',
                'inserted_count': inserted_count,
                'duplicate_items': []
            })
        elif inserted_count > 0 and len(duplicate_items) > 0:
            # ì¼ë¶€ ì„±ê³µ, ì¼ë¶€ ì¤‘ë³µ
            return jsonify({
                'success': True,
                'message': f'{inserted_count}ê°œ í•­ëª© ì‹ ì²­ ì™„ë£Œ, {len(duplicate_items)}ê°œ í•­ëª© ì¤‘ë³µ',
                'inserted_count': inserted_count,
                'duplicate_items': duplicate_items
            })
        else:
            # ëª¨ë‘ ì¤‘ë³µ
            return jsonify({
                'success': False,
                'message': 'ì„ íƒí•œ ëª¨ë“  í•­ëª©ì´ ì´ë¯¸ ì‹ ì²­ë˜ì–´ ìˆìŠµë‹ˆë‹¤',
                'inserted_count': 0,
                'duplicate_items': duplicate_items
            })
        
    except Exception as e:
        print(f"ê²€ì‚¬ì‹ ì²­ ìƒì„± ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500


@app.route('/api/inspection-requests/<int:request_id>/approve', methods=['PUT'])
@token_required
def approve_inspection_request(current_user, request_id):
    """ê²€ì‚¬ì‹ ì²­ ìŠ¹ì¸ (Level 3+ ì „ìš©)"""
    try:
        # ì‚¬ìš©ì ì •ë³´ ì¡°íšŒ
        user_info = get_user_info(current_user)
        if not user_info:
            return jsonify({'success': False, 'message': 'ì‚¬ìš©ì ì •ë³´ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        permission_level = user_info['permission_level']
        
        # Level 3+ ê¶Œí•œ í™•ì¸
        if permission_level < 3:
            return jsonify({'success': False, 'message': 'ìŠ¹ì¸ ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤'}), 403
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # í•´ë‹¹ ê²€ì‚¬ì‹ ì²­ì´ ì¡´ì¬í•˜ê³  ëŒ€ê¸°ì¤‘ì¸ì§€ í™•ì¸
        cursor.execute("""
            SELECT * FROM inspection_requests 
            WHERE id = %s AND status = 'ëŒ€ê¸°ì¤‘'
        """, (request_id,))
        
        inspection_request = cursor.fetchone()
        
        if not inspection_request:
            return jsonify({'success': False, 'message': 'ìŠ¹ì¸í•  ìˆ˜ ìˆëŠ” ê²€ì‚¬ì‹ ì²­ì´ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        # ìŠ¹ì¸ ì²˜ë¦¬
        today = datetime.datetime.now().date()
        cursor.execute("""
            UPDATE inspection_requests 
            SET status = 'ìŠ¹ì¸ë¨',
                approved_by_user_id = %s,
                approved_by_name = %s,
                approved_date = %s
            WHERE id = %s
        """, (current_user, user_info['full_name'], today, request_id))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'{inspection_request["assembly_code"]} {inspection_request["inspection_type"]} ê²€ì‚¬ì‹ ì²­ì´ ìŠ¹ì¸ë˜ì—ˆìŠµë‹ˆë‹¤'
        })
        
    except Exception as e:
        print(f"ê²€ì‚¬ì‹ ì²­ ìŠ¹ì¸ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/inspection-requests/<int:request_id>/confirm', methods=['PUT'])
@token_required
def confirm_inspection_request(current_user, request_id):
    """ê²€ì‚¬ì‹ ì²­ í™•ì • (Level 3+ ì „ìš©) - assembly_items í…Œì´ë¸”ë„ ì—…ë°ì´íŠ¸"""
    try:
        # ì‚¬ìš©ì ì •ë³´ ì¡°íšŒ
        user_info = get_user_info(current_user)
        if not user_info:
            return jsonify({'success': False, 'message': 'ì‚¬ìš©ì ì •ë³´ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        permission_level = user_info['permission_level']
        
        # Level 3+ ê¶Œí•œ í™•ì¸
        if permission_level < 3:
            return jsonify({'success': False, 'message': 'í™•ì • ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤'}), 403
        
        data = request.get_json()
        confirmed_date = data.get('confirmed_date')  # ì‹¤ì œ ê²€ì‚¬ ì™„ë£Œ ë‚ ì§œ
        
        if not confirmed_date:
            return jsonify({'success': False, 'message': 'í™•ì • ë‚ ì§œë¥¼ ì…ë ¥í•˜ì„¸ìš”'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # í•´ë‹¹ ê²€ì‚¬ì‹ ì²­ì´ ì¡´ì¬í•˜ê³  ìŠ¹ì¸ë¨ ìƒíƒœì¸ì§€ í™•ì¸
        cursor.execute("""
            SELECT * FROM inspection_requests 
            WHERE id = %s AND status = 'ìŠ¹ì¸ë¨'
        """, (request_id,))
        
        inspection_request = cursor.fetchone()
        
        if not inspection_request:
            return jsonify({'success': False, 'message': 'í™•ì •í•  ìˆ˜ ìˆëŠ” ê²€ì‚¬ì‹ ì²­ì´ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        # ê²€ì‚¬íƒ€ì…ë³„ assembly_items ì»¬ëŸ¼ ë§¤í•‘
        inspection_type_mapping = {
            'NDE': 'nde_date',
            'VIDI': 'vidi_date',
            'GALV': 'galv_date',
            'SHOT': 'shot_date',
            'PAINT': 'paint_date',
            'PACKING': 'packing_date'
        }
        
        inspection_type = inspection_request['inspection_type']
        if inspection_type not in inspection_type_mapping:
            return jsonify({'success': False, 'message': f'ì•Œ ìˆ˜ ì—†ëŠ” ê²€ì‚¬íƒ€ì…: {inspection_type}'}), 400
        
        assembly_column = inspection_type_mapping[inspection_type]
        assembly_code = inspection_request['assembly_code']
        
        # íŠ¸ëœì­ì…˜ ì‹œì‘
        try:
            # 1. inspection_requests ìƒíƒœë¥¼ í™•ì •ë¨ìœ¼ë¡œ ì—…ë°ì´íŠ¸
            cursor.execute("""
                UPDATE inspection_requests 
                SET status = 'í™•ì •ë¨',
                    confirmed_date = %s
                WHERE id = %s
            """, (confirmed_date, request_id))
            
            # 2. assembly_items í…Œì´ë¸”ì˜ í•´ë‹¹ ê³µì • ë‚ ì§œ ì—…ë°ì´íŠ¸
            update_query = f"""
                UPDATE assembly_items 
                SET {assembly_column} = %s
                WHERE assembly_code = %s
            """
            cursor.execute(update_query, (confirmed_date, assembly_code))
            
            # ì—…ë°ì´íŠ¸ëœ í–‰ì´ ìˆëŠ”ì§€ í™•ì¸
            if cursor.rowcount == 0:
                raise Exception(f'Assembly {assembly_code}ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤')
            
            connection.commit()
            cursor.close()
            connection.close()
            
            return jsonify({
                'success': True,
                'message': f'{assembly_code} {inspection_type} ê²€ì‚¬ê°€ í™•ì •ë˜ì—ˆìŠµë‹ˆë‹¤ ({confirmed_date})'
            })
            
        except Exception as e:
            connection.rollback()
            raise e
        
    except Exception as e:
        print(f"ê²€ì‚¬ì‹ ì²­ í™•ì • ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/inspection-requests/<int:request_id>', methods=['DELETE'])
@token_required
def cancel_inspection_request(current_user, request_id):
    """ê²€ì‚¬ì‹ ì²­ ì·¨ì†Œ (ê¶Œí•œë³„ ì œí•œ)"""
    try:
        # ì‚¬ìš©ì ì •ë³´ ì¡°íšŒ
        user_info = get_user_info(current_user)
        if not user_info:
            return jsonify({'success': False, 'message': 'ì‚¬ìš©ì ì •ë³´ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        permission_level = user_info['permission_level']
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # í•´ë‹¹ ê²€ì‚¬ì‹ ì²­ ì¡°íšŒ
        cursor.execute("""
            SELECT * FROM inspection_requests WHERE id = %s
        """, (request_id,))
        
        inspection_request = cursor.fetchone()
        
        if not inspection_request:
            return jsonify({'success': False, 'message': 'ê²€ì‚¬ì‹ ì²­ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        # ê¶Œí•œë³„ ì·¨ì†Œ ì œí•œ í™•ì¸
        if permission_level == 1:
            # Level 1: ë³¸ì¸ì´ ì‹ ì²­í•œ ëŒ€ê¸°ì¤‘ ìƒíƒœë§Œ ì·¨ì†Œ ê°€ëŠ¥
            if (inspection_request['requested_by_user_id'] != current_user or 
                inspection_request['status'] != 'ëŒ€ê¸°ì¤‘'):
                return jsonify({'success': False, 'message': 'ì·¨ì†Œ ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤'}), 403
        
        # Level 3+ëŠ” ëª¨ë“  ìƒíƒœ ì·¨ì†Œ ê°€ëŠ¥
        
        # íŠ¸ëœì­ì…˜ ì‹œì‘
        try:
            # í™•ì •ë¨ ìƒíƒœì—ì„œ ì·¨ì†Œí•˜ëŠ” ê²½ìš° assembly_items í…Œì´ë¸” ë¡¤ë°± í•„ìš”
            if inspection_request['status'] == 'í™•ì •ë¨':
                # ê²€ì‚¬íƒ€ì…ë³„ assembly_items ì»¬ëŸ¼ ë§¤í•‘
                inspection_type_mapping = {
                    'NDE': 'nde_date',
                    'VIDI': 'vidi_date',
                    'GALV': 'galv_date',
                    'SHOT': 'shot_date',
                    'PAINT': 'paint_date',
                    'PACKING': 'packing_date'
                }
                
                inspection_type = inspection_request['inspection_type']
                if inspection_type in inspection_type_mapping:
                    assembly_column = inspection_type_mapping[inspection_type]
                    assembly_code = inspection_request['assembly_code']
                    
                    # assembly_items í…Œì´ë¸”ì˜ í•´ë‹¹ ê³µì • ë‚ ì§œë¥¼ NULLë¡œ ë˜ëŒë¦¬ê¸°
                    update_query = f"""
                        UPDATE assembly_items 
                        SET {assembly_column} = NULL
                        WHERE assembly_code = %s
                    """
                    cursor.execute(update_query, (assembly_code,))
                    
                    # ì—…ë°ì´íŠ¸ëœ í–‰ì´ ìˆëŠ”ì§€ í™•ì¸
                    if cursor.rowcount == 0:
                        raise Exception(f'Assembly {assembly_code}ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤')
            
            # ì·¨ì†Œ ì²˜ë¦¬ (ì‹¤ì œë¡œëŠ” ìƒíƒœë§Œ ë³€ê²½)
            cursor.execute("""
                UPDATE inspection_requests 
                SET status = 'ì·¨ì†Œë¨'
                WHERE id = %s
            """, (request_id,))
            
            connection.commit()
            
        except Exception as e:
            connection.rollback()
            raise e
        cursor.close()
        connection.close()
        
        # ì„±ê³µ ë©”ì‹œì§€ ìƒì„±
        message = f'{inspection_request["assembly_code"]} {inspection_request["inspection_type"]} ê²€ì‚¬ì‹ ì²­ì´ ì·¨ì†Œë˜ì—ˆìŠµë‹ˆë‹¤'
        
        # í™•ì •ëœ í•­ëª©ì´ ì·¨ì†Œëœ ê²½ìš° ì¶”ê°€ ì •ë³´ ì œê³µ
        if inspection_request['status'] == 'í™•ì •ë¨':
            message += ' (ì¡°ë¦½í’ˆ ê³µì • ë‚ ì§œê°€ ë˜ëŒë ¤ì¡ŒìŠµë‹ˆë‹¤)'
        
        return jsonify({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        print(f"ê²€ì‚¬ì‹ ì²­ ì·¨ì†Œ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

def get_user_info(user_id):
    """ì‚¬ìš©ì ì •ë³´ ì¡°íšŒ í•¨ìˆ˜"""
    connection = get_db_connection()
    if not connection:
        return None
    
    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("SELECT id, username, full_name, permission_level, company FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
    except:
        # company ì»¬ëŸ¼ì´ ì—†ëŠ” ê²½ìš° ê¸°ë³¸ ì¿¼ë¦¬ ì‚¬ìš©
        cursor.execute("SELECT id, username, full_name, permission_level FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        if user:
            user['company'] = ''
    
    cursor.close()
    connection.close()
    
    return user

def admin_required(f):
    """Admin ê¶Œí•œ ê²€ì¦ ë°ì½”ë ˆì´í„°"""
    @wraps(f)
    def decorated(current_user, *args, **kwargs):
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT permission_level FROM users WHERE id = %s", (current_user,))
        user = cursor.fetchone()
        cursor.close()
        connection.close()
        
        if not user or user['permission_level'] < 5:
            return jsonify({'success': False, 'message': 'Admin ê¶Œí•œì´ í•„ìš”í•©ë‹ˆë‹¤'}), 403
        
        return f(current_user, *args, **kwargs)
    return decorated

@app.route('/api/admin/users', methods=['GET'])
@token_required
@admin_required
def get_users(current_user):
    """ì‚¬ìš©ì ëª©ë¡ ì¡°íšŒ (Admin ì „ìš©)"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        try:
            cursor.execute("""
                SELECT id, username, full_name, permission_level, company, is_active, created_at
                FROM users 
                ORDER BY permission_level DESC, created_at DESC
            """)
        except:
            # company ì»¬ëŸ¼ì´ ì—†ëŠ” ê²½ìš°
            cursor.execute("""
                SELECT id, username, full_name, permission_level, is_active, created_at
                FROM users 
                ORDER BY permission_level DESC, created_at DESC
            """)
        
        users = cursor.fetchall()
        cursor.close()
        connection.close()
        
        # ë‚ ì§œ í¬ë§·íŒ… ë° company ê¸°ë³¸ê°’ ì„¤ì •
        for user in users:
            if user['created_at']:
                user['created_at'] = user['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            if 'company' not in user:
                user['company'] = ''
        
        return jsonify({
            'success': True,
            'users': users
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/admin/users', methods=['POST'])
@token_required
@admin_required
def create_user(current_user):
    """ìƒˆ ì‚¬ìš©ì ìƒì„± (Admin ì „ìš©)"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password', '1234')  # ê¸°ë³¸ ë¹„ë°€ë²ˆí˜¸
        full_name = data.get('full_name')
        permission_level = data.get('permission_level', 1)
        company = data.get('company', '')
        
        # ë¹„ë°€ë²ˆí˜¸ í•´ì‹±
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        if not username or not full_name:
            return jsonify({'success': False, 'message': 'ì‚¬ìš©ìëª…ê³¼ ì´ë¦„ì€ í•„ìˆ˜ì…ë‹ˆë‹¤'}), 400
        
        # ë¹„ë°€ë²ˆí˜¸ í•´ì‹œí™”
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor()
        
        # ì¤‘ë³µ ì‚¬ìš©ìëª… í™•ì¸
        cursor.execute("SELECT id FROM users WHERE username = %s", (username,))
        if cursor.fetchone():
            return jsonify({'success': False, 'message': 'ì´ë¯¸ ì¡´ì¬í•˜ëŠ” ì‚¬ìš©ìëª…ì…ë‹ˆë‹¤'}), 400
        
        # ì‚¬ìš©ì ìƒì„±
        try:
            cursor.execute("""
                INSERT INTO users (username, password_hash, full_name, permission_level, company)
                VALUES (%s, %s, %s, %s, %s)
            """, (username, password_hash, full_name, permission_level, company))
        except:
            # company ì»¬ëŸ¼ì´ ì—†ëŠ” ê²½ìš°
            cursor.execute("""
                INSERT INTO users (username, password_hash, full_name, permission_level)
                VALUES (%s, %s, %s, %s)
            """, (username, password_hash, full_name, permission_level))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'ì‚¬ìš©ì {username}ì´ ìƒì„±ë˜ì—ˆìŠµë‹ˆë‹¤'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@token_required
@admin_required
def update_user(current_user, user_id):
    """ì‚¬ìš©ì ì •ë³´ ìˆ˜ì • (Admin ì „ìš©)"""
    try:
        data = request.get_json()
        full_name = data.get('full_name')
        permission_level = data.get('permission_level')
        company = data.get('company')
        is_active = data.get('is_active')
        password = data.get('password')
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor()
        
        # ì‚¬ìš©ì ì¡´ì¬ í™•ì¸
        cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'success': False, 'message': 'ì‚¬ìš©ìë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        # ì—…ë°ì´íŠ¸ í•„ë“œ êµ¬ì„±
        update_fields = []
        update_values = []
        
        if full_name is not None:
            update_fields.append("full_name = %s")
            update_values.append(full_name)
        
        if permission_level is not None:
            update_fields.append("permission_level = %s")
            update_values.append(permission_level)
        
        if password is not None:
            update_fields.append("password_hash = %s")
            password_hash = hashlib.sha256(password.encode()).hexdigest()
            update_values.append(password_hash)
        
        if company is not None:
            try:
                # company ì»¬ëŸ¼ì´ ìˆëŠ”ì§€ í™•ì¸
                cursor.execute("SHOW COLUMNS FROM users LIKE 'company'")
                if cursor.fetchone():
                    update_fields.append("company = %s")
                    update_values.append(company)
            except:
                pass
        
        if is_active is not None:
            update_fields.append("is_active = %s")
            update_values.append(is_active)
        
        if not update_fields:
            return jsonify({'success': False, 'message': 'ìˆ˜ì •í•  ë‚´ìš©ì´ ì—†ìŠµë‹ˆë‹¤'}), 400
        
        # ì—…ë°ì´íŠ¸ ì‹¤í–‰
        update_query = f"UPDATE users SET {', '.join(update_fields)} WHERE id = %s"
        update_values.append(user_id)
        
        cursor.execute(update_query, update_values)
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'ì‚¬ìš©ì ì •ë³´ê°€ ìˆ˜ì •ë˜ì—ˆìŠµë‹ˆë‹¤'
        })
        
    except Exception as e:
        print(f"ì‚¬ìš©ì ìˆ˜ì • ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@token_required
@admin_required
def deactivate_user(current_user, user_id):
    """ì‚¬ìš©ì ë¹„í™œì„±í™” (Admin ì „ìš©)"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # ì‚¬ìš©ì ì¡´ì¬ í™•ì¸
        cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'success': False, 'message': 'ì‚¬ìš©ìë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        # ì‚¬ìš©ì ë¹„í™œì„±í™”
        cursor.execute("UPDATE users SET is_active = FALSE WHERE id = %s", (user_id,))
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'ì‚¬ìš©ì {user["username"]}ì´ ë¹„í™œì„±í™”ë˜ì—ˆìŠµë‹ˆë‹¤'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/admin/users/<int:user_id>/delete-permanently', methods=['DELETE'])
@token_required
@admin_required
def delete_user_permanently(current_user, user_id):
    """ì‚¬ìš©ì ì™„ì „ ì‚­ì œ (Admin ì „ìš©)"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # ì‚¬ìš©ì ì¡´ì¬ í™•ì¸
        cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'success': False, 'message': 'ì‚¬ìš©ìë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        # ìê¸° ìì‹ ì„ ì‚­ì œí•˜ëŠ” ê²ƒì„ ë°©ì§€
        if user_id == current_user:
            return jsonify({'success': False, 'message': 'ìê¸° ìì‹ ì„ ì‚­ì œí•  ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 400
        
        # ì‚¬ìš©ì ì™„ì „ ì‚­ì œ
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'ì‚¬ìš©ì {user["username"]}ì´ ì™„ì „íˆ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤'
        })
        
    except Exception as e:
        print(f"ì‚¬ìš©ì ì‚­ì œ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/')
def home():
    """ë£¨íŠ¸ ê²½ë¡œ - ì„œë²„ ìƒíƒœ í™•ì¸"""
    return jsonify({
        'status': 'ok',
        'message': 'DSHI Field Pad Server is running',
        'timestamp': datetime.datetime.now().isoformat()
    })

@app.route('/api/health', methods=['GET'])
def health_check():
    """ì„œë²„ ìƒíƒœ í™•ì¸"""
    return jsonify({
        'status': 'ok',
        'message': 'DSHI Field Pad Server is running',
        'timestamp': datetime.datetime.now().isoformat()
    })

@app.route('/api/saved-list', methods=['POST'])
@token_required
def save_assembly_list(current_user):
    """ì‚¬ìš©ìë³„ ì €ì¥ëœ ë¦¬ìŠ¤íŠ¸ì— ì•„ì´í…œ ì¶”ê°€"""
    try:
        data = request.get_json()
        items = data.get('items', [])
        
        if not items:
            return jsonify({'success': False, 'message': 'ì €ì¥í•  í•­ëª©ì´ ì—†ìŠµë‹ˆë‹¤'}), 400
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # ê° í•­ëª©ì„ ì €ì¥ (ì¤‘ë³µ ì‹œ ì—…ë°ì´íŠ¸)
        saved_count = 0
        updated_count = 0
        
        for item in items:
            assembly_code = item.get('assembly_code')
            if not assembly_code:
                continue
                
            # ì¤‘ë³µ í™•ì¸
            cursor.execute("""
                SELECT id FROM user_saved_lists 
                WHERE user_id = %s AND assembly_code = %s
            """, (current_user, assembly_code))
            
            existing = cursor.fetchone()
            
            if existing:
                # ì—…ë°ì´íŠ¸
                cursor.execute("""
                    UPDATE user_saved_lists 
                    SET assembly_data = %s, updated_at = CURRENT_TIMESTAMP
                    WHERE user_id = %s AND assembly_code = %s
                """, (json.dumps(item), current_user, assembly_code))
                updated_count += 1
            else:
                # ìƒˆë¡œ ì‚½ì…
                cursor.execute("""
                    INSERT INTO user_saved_lists (user_id, assembly_code, assembly_data)
                    VALUES (%s, %s, %s)
                """, (current_user, assembly_code, json.dumps(item)))
                saved_count += 1
        
        connection.commit()
        
        # ì´ ì €ì¥ëœ í•­ëª© ìˆ˜ ì¡°íšŒ
        cursor.execute("""
            SELECT COUNT(*) as total FROM user_saved_lists WHERE user_id = %s
        """, (current_user,))
        total_result = cursor.fetchone()
        total = total_result['total'] if total_result else 0
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'{saved_count}ê°œ í•­ëª© ì¶”ê°€, {updated_count}ê°œ í•­ëª© ì—…ë°ì´íŠ¸',
            'saved_count': saved_count,
            'updated_count': updated_count,
            'total': total
        })
        
    except Exception as e:
        print(f"ì €ì¥ëœ ë¦¬ìŠ¤íŠ¸ ì¶”ê°€ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/saved-list', methods=['GET'])
@token_required
def get_saved_list(current_user):
    """ì‚¬ìš©ìë³„ ì €ì¥ëœ ë¦¬ìŠ¤íŠ¸ ì¡°íšŒ"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # ì €ì¥ëœ ë¦¬ìŠ¤íŠ¸ì™€ ì‹¤ì‹œê°„ ë°ì´í„°ë¥¼ JOINìœ¼ë¡œ í•œë²ˆì— ì¡°íšŒ (ìµœì í™”)
        cursor.execute("""
            SELECT 
                usl.assembly_code, 
                usl.created_at as saved_at, 
                usl.updated_at,
                ae.company, 
                ae.zone, 
                ae.item, 
                ae.weight_gross,
                ae.fit_up_date, 
                ae.final_date, 
                ae.arup_final_date, 
                ae.galv_date, 
                ae.arup_galv_date, 
                ae.shot_date, 
                ae.paint_date, 
                ae.arup_paint_date
            FROM user_saved_lists usl
            JOIN arup_ecs ae ON usl.assembly_code = ae.assembly_code
            WHERE usl.user_id = %s
            ORDER BY usl.updated_at DESC
        """, (current_user,))
        
        saved_items = cursor.fetchall()
        cursor.close()
        connection.close()
        
        # ì‹¤ì‹œê°„ ë°ì´í„°ë¡œ ìƒíƒœ ê³„ì‚° (JSON íŒŒì‹± ì—†ì´ ì§ì ‘ ì²˜ë¦¬)
        result_items = []
        for item in saved_items:
            try:
                # dictionary í˜•íƒœë¡œ ë³€í™˜
                assembly_data = dict(item)
                assembly_data['saved_at'] = item['saved_at'].strftime('%Y-%m-%d %H:%M:%S') if item['saved_at'] else ''
                assembly_data['updated_at'] = item['updated_at'].strftime('%Y-%m-%d %H:%M:%S') if item['updated_at'] else ''
                
                # ì‹¤ì‹œê°„ ìƒíƒœ ê³„ì‚° (ìµœì‹  ë°ì´í„° ê¸°ë°˜)
                assembly_data = calculate_assembly_status(assembly_data)
                
                result_items.append(assembly_data)
            except AttributeError as e:
                print(f"ë°ì´í„° ì²˜ë¦¬ ì˜¤ë¥˜: {e}")
                continue
        
        return jsonify({
            'success': True,
            'items': result_items,
            'total': len(result_items)
        })
        
    except Exception as e:
        print(f"ì €ì¥ëœ ë¦¬ìŠ¤íŠ¸ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/saved-list/<assembly_code>', methods=['DELETE'])
@token_required
def delete_saved_item(current_user, assembly_code):
    """ì €ì¥ëœ ë¦¬ìŠ¤íŠ¸ì—ì„œ íŠ¹ì • í•­ëª© ì‚­ì œ"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor()
        
        # í•´ë‹¹ ì‚¬ìš©ìì˜ í•­ëª©ì¸ì§€ í™•ì¸í•˜ê³  ì‚­ì œ
        cursor.execute("""
            DELETE FROM user_saved_lists 
            WHERE user_id = %s AND assembly_code = %s
        """, (current_user, assembly_code))
        
        if cursor.rowcount == 0:
            return jsonify({'success': False, 'message': 'ì‚­ì œí•  í•­ëª©ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'{assembly_code} í•­ëª©ì´ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤'
        })
        
    except Exception as e:
        print(f"ì €ì¥ëœ í•­ëª© ì‚­ì œ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/saved-list/clear', methods=['DELETE'])
@token_required
def clear_saved_list(current_user):
    """ì‚¬ìš©ìì˜ ì €ì¥ëœ ë¦¬ìŠ¤íŠ¸ ì „ì²´ ì‚­ì œ"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor()
        
        # í•´ë‹¹ ì‚¬ìš©ìì˜ ëª¨ë“  ì €ì¥ í•­ëª© ì‚­ì œ
        cursor.execute("""
            DELETE FROM user_saved_lists WHERE user_id = %s
        """, (current_user,))
        
        deleted_count = cursor.rowcount
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'{deleted_count}ê°œ í•­ëª©ì´ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        print(f"ì €ì¥ëœ ë¦¬ìŠ¤íŠ¸ ì „ì²´ ì‚­ì œ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

# =================================
# ê²€ì‚¬ì‹ ì²­ ê´€ë¦¬ API 
# =================================

@app.route('/api/inspection-management/requests', methods=['GET'])
@token_required
def get_inspection_management_requests(current_user):
    """ê²€ì‚¬ì‹ ì²­ ëª©ë¡ ì¡°íšŒ"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # ì‚¬ìš©ì ì •ë³´ ì¡°íšŒ
        cursor.execute("SELECT id, username, permission_level FROM users WHERE id = %s", (current_user,))
        user_info = cursor.fetchone()
        
        if not user_info:
            return jsonify({'success': False, 'message': 'ì‚¬ìš©ì ì •ë³´ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        # Level 1 ì‚¬ìš©ìëŠ” ë³¸ì¸ ì‹ ì²­ê±´ë§Œ, Level 2+ ì‚¬ìš©ìëŠ” ì „ì²´ ì¡°íšŒ
        if user_info['permission_level'] == 1:
            # Level 1: ë³¸ì¸ ì‹ ì²­ê±´ë§Œ + í™•ì •ë˜ì§€ ì•Šì€ ê±´ê³¼ ì·¨ì†Œëœ ê±´ ì œì™¸ (í•œê¸€ ìƒíƒœê°’ ì‚¬ìš©)
            query = """
                SELECT ir.*, u.username as requested_by_name
                FROM inspection_requests ir
                LEFT JOIN users u ON ir.requested_by_user_id = u.id
                WHERE ir.requested_by_user_id = %s 
                AND ir.status NOT IN ('í™•ì •ë¨', 'ì·¨ì†Œë¨')
                ORDER BY ir.created_at DESC
            """
            cursor.execute(query, (current_user,))
        else:
            # Level 2+: ì „ì²´ ì¡°íšŒ
            query = """
                SELECT ir.*, u.username as requested_by_name
                FROM inspection_requests ir
                LEFT JOIN users u ON ir.requested_by_user_id = u.id
                ORDER BY ir.created_at DESC
            """
            cursor.execute(query)
        
        requests = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': {
                'requests': requests,
                'user_level': user_info['permission_level']
            }
        })
        
    except Exception as e:
        print(f"ê²€ì‚¬ì‹ ì²­ ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/inspection-management/requests/<int:request_id>/approve', methods=['PUT'])
@token_required
def approve_inspection_management_request(current_user, request_id):
    """ê²€ì‚¬ì‹ ì²­ ìŠ¹ì¸ (Level 2+ ê¶Œí•œ í•„ìš”)"""
    logger.debug(f"ê²€ì‚¬ì‹ ì²­ ìŠ¹ì¸ ìš”ì²­: request_id={request_id}, user_id={current_user}")
    try:
        # ì‚¬ìš©ì ê¶Œí•œ í™•ì¸
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # ìŠ¹ì¸ì ì •ë³´ ì¡°íšŒ
        cursor.execute("SELECT id, username, permission_level FROM users WHERE id = %s", (current_user,))
        approver = cursor.fetchone()
        
        if not approver or approver['permission_level'] < 2:
            return jsonify({'success': False, 'message': 'ìŠ¹ì¸ ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤ (Level 2+ í•„ìš”)'}), 403
        
        # ê²€ì‚¬ì‹ ì²­ ì¡´ì¬ ë° ìƒíƒœ í™•ì¸
        cursor.execute("SELECT * FROM inspection_requests WHERE id = %s", (request_id,))
        request_data = cursor.fetchone()
        
        if not request_data:
            logger.debug(f"ê²€ì‚¬ì‹ ì²­ì„ ì°¾ì„ ìˆ˜ ì—†ìŒ: request_id={request_id}")
            return jsonify({'success': False, 'message': 'ê²€ì‚¬ì‹ ì²­ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        logger.debug(f"ê²€ì‚¬ì‹ ì²­ í˜„ì¬ ìƒíƒœ: {request_data['status']}")
        if request_data['status'] not in ['pending', 'ëŒ€ê¸°ì¤‘']:
            logger.debug(f"ìŠ¹ì¸ ë¶ˆê°€ ìƒíƒœ: {request_data['status']}")
            return jsonify({'success': False, 'message': 'ëŒ€ê¸°ì¤‘ì¸ ê²€ì‚¬ì‹ ì²­ë§Œ ìŠ¹ì¸í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤'}), 400
        
        # ìŠ¹ì¸ ì²˜ë¦¬
        cursor.execute("""
            UPDATE inspection_requests 
            SET status = 'ìŠ¹ì¸ë¨',
                approved_by = %s,
                approved_by_name = %s,
                approved_date = CURRENT_DATE,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (current_user, approver['username'], request_id))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        logger.debug(f"ê²€ì‚¬ì‹ ì²­ ìŠ¹ì¸ ì™„ë£Œ: request_id={request_id}, approver={approver['username']}")
        return jsonify({
            'success': True,
            'message': f'ê²€ì‚¬ì‹ ì²­ì´ ìŠ¹ì¸ë˜ì—ˆìŠµë‹ˆë‹¤ (ìŠ¹ì¸ì: {approver["username"]})'
        })
        
    except Exception as e:
        logger.error(f"ê²€ì‚¬ì‹ ì²­ ìŠ¹ì¸ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/inspection-management/requests/<int:request_id>/reject', methods=['PUT'])
@token_required
def reject_inspection_management_request(current_user, request_id):
    """ê²€ì‚¬ì‹ ì²­ ê±°ë¶€ (Level 2+ ê¶Œí•œ í•„ìš”)"""
    try:
        data = request.get_json()
        reject_reason = data.get('reject_reason', 'ê±°ë¶€ë¨')
        
        # ì‚¬ìš©ì ê¶Œí•œ í™•ì¸
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # ê±°ë¶€ì ì •ë³´ ì¡°íšŒ
        cursor.execute("SELECT id, username, permission_level FROM users WHERE id = %s", (current_user,))
        rejecter = cursor.fetchone()
        
        if not rejecter or rejecter['permission_level'] < 2:
            return jsonify({'success': False, 'message': 'ê±°ë¶€ ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤ (Level 2+ í•„ìš”)'}), 403
        
        # ê²€ì‚¬ì‹ ì²­ ì¡´ì¬ ë° ìƒíƒœ í™•ì¸
        cursor.execute("SELECT * FROM inspection_requests WHERE id = %s", (request_id,))
        request_data = cursor.fetchone()
        
        if not request_data:
            return jsonify({'success': False, 'message': 'ê²€ì‚¬ì‹ ì²­ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        if request_data['status'] not in ['pending', 'ëŒ€ê¸°ì¤‘', 'approved', 'ìŠ¹ì¸ë¨']:
            return jsonify({'success': False, 'message': 'ëŒ€ê¸°ì¤‘ì´ê±°ë‚˜ ìŠ¹ì¸ëœ ê²€ì‚¬ì‹ ì²­ë§Œ ê±°ë¶€í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤'}), 400
        
        # ê±°ë¶€ ì²˜ë¦¬
        cursor.execute("""
            UPDATE inspection_requests 
            SET status = 'ê±°ë¶€ë¨',
                reject_reason = %s,
                approved_by = %s,
                approved_by_name = %s,
                approved_date = CURRENT_DATE,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (reject_reason, current_user, rejecter['username'], request_id))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'ê²€ì‚¬ì‹ ì²­ì´ ê±°ë¶€ë˜ì—ˆìŠµë‹ˆë‹¤ (ê±°ë¶€ì: {rejecter["username"]})'
        })
        
    except Exception as e:
        print(f"ê²€ì‚¬ì‹ ì²­ ê±°ë¶€ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/inspection-management/requests/<int:request_id>/confirm', methods=['PUT'])
@token_required
def confirm_inspection_management_request(current_user, request_id):
    """ê²€ì‚¬ì‹ ì²­ í™•ì • (Level 3+ ê¶Œí•œ í•„ìš”)"""
    try:
        data = request.get_json()
        confirmed_date = data.get('confirmed_date')
        
        if not confirmed_date:
            return jsonify({'success': False, 'message': 'í™•ì • ë‚ ì§œë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”'}), 400
        
        # ì‚¬ìš©ì ê¶Œí•œ í™•ì¸
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # í™•ì •ì ì •ë³´ ì¡°íšŒ
        cursor.execute("SELECT id, username, permission_level FROM users WHERE id = %s", (current_user,))
        confirmer = cursor.fetchone()
        
        if not confirmer or confirmer['permission_level'] < 3:
            return jsonify({'success': False, 'message': 'í™•ì • ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤ (Level 3+ í•„ìš”)'}), 403
        
        # ê²€ì‚¬ì‹ ì²­ ì¡´ì¬ ë° ìƒíƒœ í™•ì¸
        cursor.execute("SELECT * FROM inspection_requests WHERE id = %s", (request_id,))
        request_data = cursor.fetchone()
        
        if not request_data:
            return jsonify({'success': False, 'message': 'ê²€ì‚¬ì‹ ì²­ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        if request_data['status'] not in ['approved', 'ìŠ¹ì¸ë¨']:
            return jsonify({'success': False, 'message': 'ìŠ¹ì¸ëœ ê²€ì‚¬ì‹ ì²­ë§Œ í™•ì •í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤'}), 400
        
        # í™•ì • ì²˜ë¦¬
        cursor.execute("""
            UPDATE inspection_requests 
            SET status = 'í™•ì •ë¨',
                confirmed_by = %s,
                confirmed_by_name = %s,
                confirmed_date = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (current_user, confirmer['username'], confirmed_date, request_id))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'ê²€ì‚¬ì‹ ì²­ì´ í™•ì •ë˜ì—ˆìŠµë‹ˆë‹¤ (í™•ì •ì: {confirmer["username"]})'
        })
        
    except Exception as e:
        print(f"ê²€ì‚¬ì‹ ì²­ í™•ì • ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/inspection-management/requests/<int:request_id>/cancel', methods=['PUT'])
@token_required
def cancel_inspection_management_request(current_user, request_id):
    """ê²€ì‚¬ì‹ ì²­ ì·¨ì†Œ (ë³¸ì¸ ì‹ ì²­ê±´ë§Œ, ëŒ€ê¸°ì¤‘ ìƒíƒœë§Œ)"""
    try:
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # ì‚¬ìš©ì ì •ë³´ ì¡°íšŒ
        cursor.execute("SELECT id, username FROM users WHERE id = %s", (current_user,))
        user_info = cursor.fetchone()
        
        if not user_info:
            return jsonify({'success': False, 'message': 'ì‚¬ìš©ì ì •ë³´ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        # ê²€ì‚¬ì‹ ì²­ ì¡´ì¬ ë° ê¶Œí•œ í™•ì¸
        cursor.execute("SELECT * FROM inspection_requests WHERE id = %s", (request_id,))
        request_data = cursor.fetchone()
        
        if not request_data:
            return jsonify({'success': False, 'message': 'ê²€ì‚¬ì‹ ì²­ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        # ë³¸ì¸ ì‹ ì²­ê±´ì¸ì§€ í™•ì¸
        if request_data['requested_by_user_id'] != current_user:
            return jsonify({'success': False, 'message': 'ë³¸ì¸ì´ ì‹ ì²­í•œ ê²€ì‚¬ì‹ ì²­ë§Œ ì·¨ì†Œí•  ìˆ˜ ìˆìŠµë‹ˆë‹¤'}), 403
        
        # ëŒ€ê¸°ì¤‘ ìƒíƒœì¸ì§€ í™•ì¸
        if request_data['status'] not in ['pending', 'ëŒ€ê¸°ì¤‘']:
            return jsonify({'success': False, 'message': 'ëŒ€ê¸°ì¤‘ì¸ ê²€ì‚¬ì‹ ì²­ë§Œ ì·¨ì†Œí•  ìˆ˜ ìˆìŠµë‹ˆë‹¤'}), 400
        
        # ì·¨ì†Œ ì²˜ë¦¬
        cursor.execute("""
            UPDATE inspection_requests 
            SET status = 'ì·¨ì†Œë¨',
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (request_id,))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'ê²€ì‚¬ì‹ ì²­ì´ ì·¨ì†Œë˜ì—ˆìŠµë‹ˆë‹¤ (ì‹ ì²­ì: {user_info["username"]})'
        })
        
    except Exception as e:
        print(f"ê²€ì‚¬ì‹ ì²­ ì·¨ì†Œ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

@app.route('/api/inspection-management/requests/<int:request_id>', methods=['DELETE'])
@token_required
def delete_inspection_management_request(current_user, request_id):
    """ê²€ì‚¬ì‹ ì²­ ì‚­ì œ (Level 3+ ê¶Œí•œ í•„ìš”)"""
    try:
        # ì‚¬ìš©ì ê¶Œí•œ í™•ì¸
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # ì‚­ì œì ì •ë³´ ì¡°íšŒ
        cursor.execute("SELECT id, username, permission_level FROM users WHERE id = %s", (current_user,))
        deleter = cursor.fetchone()
        
        if not deleter or deleter['permission_level'] < 3:
            return jsonify({'success': False, 'message': 'ì‚­ì œ ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤ (Level 3+ í•„ìš”)'}), 403
        
        # ê²€ì‚¬ì‹ ì²­ ì¡´ì¬ í™•ì¸
        cursor.execute("SELECT assembly_code FROM inspection_requests WHERE id = %s", (request_id,))
        request_data = cursor.fetchone()
        
        if not request_data:
            return jsonify({'success': False, 'message': 'ê²€ì‚¬ì‹ ì²­ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 404
        
        # ì‚­ì œ ì²˜ë¦¬
        cursor.execute("DELETE FROM inspection_requests WHERE id = %s", (request_id,))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': f'ê²€ì‚¬ì‹ ì²­ì´ ì‚­ì œë˜ì—ˆìŠµë‹ˆë‹¤ (ì¡°ë¦½í’ˆ: {request_data["assembly_code"]}, ì‚­ì œì: {deleter["username"]})'
        })
        
    except Exception as e:
        print(f"ê²€ì‚¬ì‹ ì²­ ì‚­ì œ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

# =================================
# Dashboard API (Level 3+ ì „ìš©)
# =================================


@app.route('/api/dashboard-data', methods=['GET'])
@token_required
def get_dashboard_data(current_user):
    """ëŒ€ì‹œë³´ë“œ ë°ì´í„° ì œê³µ API (Level 3+ ê¶Œí•œ í•„ìš”)"""
    try:
        # ì‚¬ìš©ì ê¶Œí•œ í™•ì¸
        user_info = get_user_info(current_user)
        if not user_info or user_info['permission_level'] < 3:
            return jsonify({'success': False, 'message': 'ëŒ€ì‹œë³´ë“œ ì ‘ê·¼ ê¶Œí•œì´ ì—†ìŠµë‹ˆë‹¤ (Level 3+ í•„ìš”)'}), 403
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'success': False, 'message': 'ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²° ì‹¤íŒ¨'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # 1. ì „ì²´ í†µê³„
        cursor.execute("""
            SELECT 
                COUNT(*) as total_assemblies,
                SUM(weight_gross) as total_weight
            FROM arup_ecs 
            WHERE weight_gross IS NOT NULL
        """)
        overall_stats = cursor.fetchone()
        
        # 2. 8ë‹¨ê³„ ê³µì •ë³„ ì™„ë£Œìœ¨ (ì¤‘ëŸ‰ ê¸°ì¤€)
        processes = [
            ('FIT_UP', 'fit_up_date'),
            ('FINAL', 'final_date'),
            ('ARUP_FINAL', 'arup_final_date'),
            ('GALV', 'galv_date'),
            ('ARUP_GALV', 'arup_galv_date'),
            ('SHOT', 'shot_date'),
            ('PAINT', 'paint_date'),
            ('ARUP_PAINT', 'arup_paint_date')
        ]
        
        process_completion = {}
        for process_name, date_column in processes:
            cursor.execute(f"""
                SELECT COALESCE(SUM(weight_gross), 0) as completed_weight
                FROM arup_ecs 
                WHERE {date_column} IS NOT NULL 
                AND {date_column} != '1900-01-01'
                AND weight_gross IS NOT NULL
            """)
            result = cursor.fetchone()
            completed_weight = result['completed_weight']
            percentage = round((completed_weight / overall_stats['total_weight']) * 100, 1) if overall_stats['total_weight'] > 0 else 0
            process_completion[process_name] = percentage
        
        # 3. ITEMë³„ ê³µì •ë¥  (BEAM/POST íƒ­)
        item_process_completion = {}
        for item_type in ['BEAM', 'POST']:
            item_process_completion[item_type] = {}
            
            # í•´ë‹¹ ITEMì˜ ì´ ì¤‘ëŸ‰ ê³„ì‚°
            cursor.execute("""
                SELECT COALESCE(SUM(weight_gross), 0) as item_total_weight
                FROM arup_ecs 
                WHERE item = %s AND weight_gross IS NOT NULL
            """, (item_type,))
            item_total = cursor.fetchone()['item_total_weight']
            
            # ê° ê³µì •ë³„ ì™„ë£Œìœ¨ ê³„ì‚°
            for process_name, date_column in processes:
                cursor.execute(f"""
                    SELECT COALESCE(SUM(weight_gross), 0) as completed_weight
                    FROM arup_ecs 
                    WHERE item = %s 
                    AND {date_column} IS NOT NULL 
                    AND {date_column} != '1900-01-01'
                    AND weight_gross IS NOT NULL
                """, (item_type,))
                result = cursor.fetchone()
                completed_weight = result['completed_weight']
                percentage = round((completed_weight / item_total) * 100, 1) if item_total > 0 else 0
                item_process_completion[item_type][process_name] = percentage
        
        # 4. ì—…ì²´ë³„ ë¶„í¬ (ì¤‘ëŸ‰ ê¸°ì¤€)
        cursor.execute("""
            SELECT 
                company,
                COUNT(*) as count,
                SUM(weight_gross) as total_weight,
                ROUND((SUM(weight_gross) / %s) * 100, 1) as percentage
            FROM arup_ecs 
            WHERE weight_gross IS NOT NULL AND company IS NOT NULL
            GROUP BY company
            ORDER BY total_weight DESC
        """, (overall_stats['total_weight'],))
        company_distribution = cursor.fetchall()
        
        # 5. ì „ì²´ ì§„í–‰ë¥  ê³„ì‚° (ARUP_PAINT ì™„ë£Œëœ ì¤‘ëŸ‰ ê¸°ì¤€)
        cursor.execute("""
            SELECT COALESCE(SUM(weight_gross), 0) as completed_weight
            FROM arup_ecs 
            WHERE arup_paint_date IS NOT NULL 
            AND arup_paint_date != '1900-01-01'
            AND weight_gross IS NOT NULL
        """)
        completed_total = cursor.fetchone()['completed_weight']
        overall_progress = round((completed_total / overall_stats['total_weight']) * 100, 1) if overall_stats['total_weight'] > 0 else 0
        
        cursor.close()
        connection.close()
        
        # 6. ë°ì´í„° ì†ŒìŠ¤ ì •ë³´
        data_source = {
            'updated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_assemblies': overall_stats['total_assemblies'],
            'total_weight_tons': round(overall_stats['total_weight'] / 1000, 1) if overall_stats['total_weight'] else 0
        }
        
        # ìµœì¢… ëŒ€ì‹œë³´ë“œ ë°ì´í„° êµ¬ì„±
        dashboard_data = {
            'overall_stats': {
                'total_assemblies': overall_stats['total_assemblies'],
                'total_weight': round(overall_stats['total_weight'], 1) if overall_stats['total_weight'] else 0,
                'total_weight_tons': data_source['total_weight_tons'],
                'overall_progress': overall_progress
            },
            'process_completion': process_completion,
            'item_process_completion': item_process_completion,
            'company_distribution': company_distribution,
            'data_source': data_source
        }
        
        return jsonify({
            'success': True,
            'data': dashboard_data
        })
        
    except Exception as e:
        print(f"Dashboard ë°ì´í„° ì¡°íšŒ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

# =================================
# Excel ì—…ë¡œë“œ API
# =================================

@app.route('/api/upload-excel', methods=['POST', 'OPTIONS'])
@token_required
def upload_excel(current_user):
    """Excel íŒŒì¼ ì—…ë¡œë“œ ë° Assembly Code íŒŒì‹±"""
    try:
        # OPTIONS ìš”ì²­ ì²˜ë¦¬ (CORS)
        if request.method == 'OPTIONS':
            response = jsonify({'success': True})
            response.headers['Access-Control-Allow-Origin'] = 'http://localhost:5008'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
            response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            return response
        
        logger.debug(f"Excel ì—…ë¡œë“œ ìš”ì²­ ì‹œì‘ - ì‚¬ìš©ì: {current_user}")
        
        # íŒŒì¼ í™•ì¸
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'íŒŒì¼ì´ ì—…ë¡œë“œë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'íŒŒì¼ì´ ì„ íƒë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤'}), 400
        
        # íŒŒì¼ í™•ì¥ì í™•ì¸
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Excel íŒŒì¼(.xlsx, .xls)ë§Œ ì—…ë¡œë“œ ê°€ëŠ¥í•©ë‹ˆë‹¤'}), 400
        
        logger.debug(f"ì—…ë¡œë“œëœ íŒŒì¼: {file.filename}")
        
        # ì„ì‹œ íŒŒì¼ë¡œ ì €ì¥
        filename = secure_filename(file.filename)
        temp_path = os.path.join('temp', filename)
        os.makedirs('temp', exist_ok=True)
        file.save(temp_path)
        
        try:
            # Excel íŒŒì¼ íŒŒì‹±
            workbook = load_workbook(temp_path, read_only=True)
            sheet = workbook.active
            
            assembly_codes = []
            for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_col=1, values_only=True), 1):
                if row_num > 100:  # ìµœëŒ€ 100ê°œ ì œí•œ
                    break
                    
                cell_value = row[0]
                if cell_value and str(cell_value).strip():
                    assembly_codes.append(str(cell_value).strip())
            
            workbook.close()
            
            logger.debug(f"íŒŒì‹±ëœ Assembly Code ìˆ˜: {len(assembly_codes)}")
            
            if not assembly_codes:
                return jsonify({'success': False, 'message': 'Aì—´ì—ì„œ Assembly Codeë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤'}), 400
            
            # íŒŒì‹±ëœ ë°ì´í„°ë¥¼ upload_assembly_codes í•¨ìˆ˜ë¡œ ì „ë‹¬
            request_data = {'assembly_codes': assembly_codes}
            
            # ë‚´ë¶€ì ìœ¼ë¡œ upload_assembly_codes í˜¸ì¶œ
            return upload_assembly_codes_internal(current_user, assembly_codes)
            
        finally:
            # ì„ì‹œ íŒŒì¼ ì‚­ì œ
            if os.path.exists(temp_path):
                os.remove(temp_path)
        
        # ì„±ê³µ ì‘ë‹µì— CORS í—¤ë” ì¶”ê°€
        response = jsonify({'success': True, 'message': 'Excel íŒŒì¼ ì—…ë¡œë“œ ì„±ê³µ'})
        response.headers['Access-Control-Allow-Origin'] = 'http://localhost:5008'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        return response
        
    except Exception as e:
        logger.error(f"Excel ì—…ë¡œë“œ ì˜¤ë¥˜: {e}")
        error_response = jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'})
        # ì—ëŸ¬ ì‘ë‹µì—ë„ CORS í—¤ë” ì¶”ê°€
        error_response.headers['Access-Control-Allow-Origin'] = 'http://localhost:5008'
        error_response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        error_response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        return error_response, 500

@app.route('/api/upload-assembly-codes', methods=['POST'])
@token_required
def upload_assembly_codes(current_user):
    """Assembly Code ëª©ë¡ ì—…ë¡œë“œ ë° ì €ì¥ë¦¬ìŠ¤íŠ¸ ì¶”ê°€ (íŒŒì¼ ì—†ì´ ë°ì´í„°ë§Œ)"""
    try:
        # JSON ë°ì´í„° í™•ì¸
        if not request.json or 'assembly_codes' not in request.json:
            return jsonify({'success': False, 'message': 'Assembly Code ëª©ë¡ì´ ì—†ìŠµë‹ˆë‹¤'}), 400
        
        assembly_codes = request.json['assembly_codes']
        return upload_assembly_codes_internal(current_user, assembly_codes)
        
    except Exception as e:
        logger.error(f"Assembly Code ì—…ë¡œë“œ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

def upload_assembly_codes_internal(current_user, assembly_codes):
    """Assembly Code ëª©ë¡ ì²˜ë¦¬ ë‚´ë¶€ í•¨ìˆ˜"""
    try:
        if not assembly_codes or not isinstance(assembly_codes, list):
            return jsonify({'success': False, 'message': 'Assembly Code ëª©ë¡ì´ ë¹„ì–´ìˆìŠµë‹ˆë‹¤'}), 400
        
        # 100ê°œ ì œí•œ í™•ì¸
        if len(assembly_codes) > 100:
            return jsonify({
                'success': False, 
                'message': f'ìµœëŒ€ 100ê°œê¹Œì§€ë§Œ ì—…ë¡œë“œ ê°€ëŠ¥í•©ë‹ˆë‹¤. í˜„ì¬: {len(assembly_codes)}ê°œ'
            }), 400
        
        logger.debug(f"Assembly Code ëª©ë¡ ì—…ë¡œë“œ ì‹œì‘: {len(assembly_codes)}ê°œ")
        
        # ë°ì´í„°ë² ì´ìŠ¤ ì—°ê²°
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # ê° Assembly Code ì¡°íšŒ ë° ìœ íš¨ì„± ê²€ì‚¬
        valid_assemblies = []
        invalid_codes = []
        
        for code in assembly_codes:
            if not code or not str(code).strip():
                continue
                
            code = str(code).strip()
            
            try:
                cursor.execute("""
                    SELECT assembly_code, company, zone, item, weight_gross,
                           fit_up_date, final_date, arup_final_date, galv_date,
                           arup_galv_date, shot_date, paint_date, arup_paint_date
                    FROM arup_ecs 
                    WHERE assembly_code = %s
                """, (code,))
                
                result = cursor.fetchone()
                if result:
                    # ì¡°ë¦½í’ˆ ë°ì´í„° êµ¬ì„± (ì‹¤ì œ arup_ecs í…Œì´ë¸” êµ¬ì¡°ì— ë§ê²Œ)
                    assembly_data = {
                        'assembly_code': result[0],
                        'company': result[1] or '',
                        'zone': result[2] or '',
                        'item': result[3] or '',
                        'weight_net': float(result[4]) if result[4] else 0.0,  # í˜¸í™˜ì„±ì„ ìœ„í•´ weight_netìœ¼ë¡œ ìœ ì§€
                        'fit_up_date': result[5],
                        'final_date': result[6],
                        'arup_final_date': result[7],
                        'galv_date': result[8],
                        'arup_galv_date': result[9],
                        'shot_date': result[10],
                        'paint_date': result[11],
                        'arup_paint_date': result[12]
                    }
                    
                    # ìƒíƒœ ê³„ì‚° í›„ ì¶”ê°€
                    assembly_data = calculate_assembly_status(assembly_data)
                    valid_assemblies.append(assembly_data)
                else:
                    invalid_codes.append(code)
                    
            except Exception as e:
                logger.error(f"Assembly Code ì¡°íšŒ ì˜¤ë¥˜ ({code}): {e}")
                invalid_codes.append(code)
        
        logger.debug(f"ìœ íš¨í•œ Assembly: {len(valid_assemblies)}ê°œ, ë¬´íš¨í•œ ì½”ë“œ: {len(invalid_codes)}ê°œ")
        
        # ìœ íš¨í•œ Assemblyê°€ ìˆìœ¼ë©´ ì €ì¥ë¦¬ìŠ¤íŠ¸ì— ì¶”ê°€
        saved_count = 0
        updated_count = 0
        
        # JWT ë°ì´í„°ì—ì„œ user_idì™€ username ì¶”ì¶œ
        user_id = current_user.get('user_id')
        username = current_user.get('username')
        logger.debug(f"JWTì—ì„œ ì¶”ì¶œí•œ user_id: {user_id}, username: {username}")
        
        if valid_assemblies and username and user_id:
            for assembly in valid_assemblies:
                try:
                    assembly_code = assembly['assembly_code']
                    
                    # ë‚ ì§œ í•„ë“œë“¤ì„ ëª…ì‹œì ìœ¼ë¡œ ë¬¸ìì—´ë¡œ ë³€í™˜ (JSON ì§ë ¬í™” ì˜¤ë¥˜ ë°©ì§€)
                    assembly_safe = dict(assembly)
                    date_fields = ['fit_up_date', 'final_date', 'arup_final_date', 'galv_date', 
                                   'arup_galv_date', 'shot_date', 'paint_date', 'arup_paint_date']
                    
                    for field in date_fields:
                        if field in assembly_safe and assembly_safe[field] is not None:
                            assembly_safe[field] = str(assembly_safe[field])
                    
                    # ì´ë¯¸ ì¡´ì¬í•˜ëŠ”ì§€ í™•ì¸ (user_id ì‚¬ìš©)
                    cursor.execute("""
                        SELECT id FROM user_saved_lists 
                        WHERE user_id = %s AND assembly_code = %s
                    """, (user_id, assembly_code))
                    
                    existing = cursor.fetchone()
                    
                    if existing:
                        # ì—…ë°ì´íŠ¸ (ë®ì–´ì“°ê¸°)
                        cursor.execute("""
                            UPDATE user_saved_lists 
                            SET assembly_data = %s, updated_at = CURRENT_TIMESTAMP
                            WHERE user_id = %s AND assembly_code = %s
                        """, (json.dumps(assembly_safe), user_id, assembly_code))
                        updated_count += 1
                    else:
                        # ìƒˆë¡œ ì‚½ì…
                        cursor.execute("""
                            INSERT INTO user_saved_lists (user_id, assembly_code, assembly_data)
                            VALUES (%s, %s, %s)
                        """, (user_id, assembly_code, json.dumps(assembly_safe)))
                        saved_count += 1
                        
                except Exception as e:
                    logger.error(f"ì €ì¥ë¦¬ìŠ¤íŠ¸ ì¶”ê°€ ì˜¤ë¥˜ ({assembly['assembly_code']}): {e}")
        
        connection.commit()
        connection.close()
        
        # ê²°ê³¼ ë°ì´í„° êµ¬ì„±
        result_data = {
            'total_uploaded': len(assembly_codes),
            'valid_count': len(valid_assemblies),
            'invalid_count': len(invalid_codes),
            'saved_count': saved_count,
            'updated_count': updated_count,
            'total_in_list': saved_count + updated_count,
            'invalid_codes': invalid_codes if invalid_codes else None
        }
        
        logger.debug(f"Assembly Code ì—…ë¡œë“œ ì™„ë£Œ: {result_data}")
        
        return jsonify({
            'success': True,
            'message': f'ì—…ë¡œë“œ ì™„ë£Œ! ìœ íš¨: {len(valid_assemblies)}ê°œ, ë¬´íš¨: {len(invalid_codes)}ê°œ',
            'data': result_data
        })
        
    except Exception as e:
        logger.error(f"Assembly Code ì²˜ë¦¬ ì˜¤ë¥˜: {e}")
        return jsonify({'success': False, 'message': f'ì„œë²„ ì˜¤ë¥˜: {str(e)}'}), 500

if __name__ == '__main__':
    print("DSHI Field Pad Server starting...")
    print(f"Server URL: http://{SERVER_CONFIG['host']}:{SERVER_CONFIG['port']}")
    app.run(**SERVER_CONFIG)