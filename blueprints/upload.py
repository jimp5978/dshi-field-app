"""
업로드 관련 블루프린트
"""
import os
import json
import logging
from flask import Blueprint, request, jsonify
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from utils.database import get_db_connection
from utils.auth_utils import token_required
from utils.assembly_utils import calculate_assembly_status

upload_bp = Blueprint('upload', __name__)
logger = logging.getLogger(__name__)

@upload_bp.route('/upload-excel', methods=['POST', 'OPTIONS'])
@token_required
def upload_excel(current_user):
    """Excel 파일 업로드 및 Assembly Code 파싱"""
    try:
        # OPTIONS 요청 처리 (CORS)
        if request.method == 'OPTIONS':
            response = jsonify({'success': True})
            response.headers['Access-Control-Allow-Origin'] = 'http://localhost:5008'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
            response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
            return response
        
        logger.debug(f"Excel 업로드 요청 시작 - 사용자: {current_user}")
        
        # 파일 확인
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '파일이 업로드되지 않았습니다'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '파일이 선택되지 않았습니다'}), 400
        
        # 파일 확장자 확인
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Excel 파일(.xlsx, .xls)만 업로드 가능합니다'}), 400
        
        logger.debug(f"업로드된 파일: {file.filename}")
        
        # 임시 파일로 저장
        filename = secure_filename(file.filename)
        temp_path = os.path.join('temp', filename)
        os.makedirs('temp', exist_ok=True)
        file.save(temp_path)
        
        try:
            # Excel 파일 파싱
            workbook = load_workbook(temp_path, read_only=True)
            sheet = workbook.active
            
            assembly_codes = []
            for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_col=1, values_only=True), 1):
                if row_num > 100:  # 최대 100개 제한
                    break
                    
                cell_value = row[0]
                if cell_value and str(cell_value).strip():
                    assembly_codes.append(str(cell_value).strip())
            
            workbook.close()
            
            logger.debug(f"파싱된 Assembly Code 수: {len(assembly_codes)}")
            
            if not assembly_codes:
                return jsonify({'success': False, 'message': 'A열에서 Assembly Code를 찾을 수 없습니다'}), 400
            
            # 파싱된 데이터를 upload_assembly_codes 함수로 전달
            request_data = {'assembly_codes': assembly_codes}
            
            # 내부적으로 upload_assembly_codes 호출
            return upload_assembly_codes_internal(current_user, assembly_codes)
            
        finally:
            # 임시 파일 삭제
            if os.path.exists(temp_path):
                os.remove(temp_path)
        
        # 성공 응답에 CORS 헤더 추가
        response = jsonify({'success': True, 'message': 'Excel 파일 업로드 성공'})
        response.headers['Access-Control-Allow-Origin'] = 'http://localhost:5008'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        return response
        
    except Exception as e:
        logger.error(f"Excel 업로드 오류: {e}")
        error_response = jsonify({'success': False, 'message': f'서버 오류: {str(e)}'})
        # 에러 응답에도 CORS 헤더 추가
        error_response.headers['Access-Control-Allow-Origin'] = 'http://localhost:5008'
        error_response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        error_response.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        return error_response, 500

@upload_bp.route('/upload-assembly-codes', methods=['POST'])
@token_required
def upload_assembly_codes(current_user):
    """Assembly Code 목록 업로드 및 저장리스트 추가 (파일 없이 데이터만)"""
    try:
        # JSON 데이터 확인
        if not request.json or 'assembly_codes' not in request.json:
            return jsonify({'success': False, 'message': 'Assembly Code 목록이 없습니다'}), 400
        
        assembly_codes = request.json['assembly_codes']
        return upload_assembly_codes_internal(current_user, assembly_codes)
        
    except Exception as e:
        logger.error(f"Assembly Code 업로드 오류: {e}")
        return jsonify({'success': False, 'message': f'서버 오류: {str(e)}'}), 500

def upload_assembly_codes_internal(current_user, assembly_codes):
    """Assembly Code 목록 처리 내부 함수"""
    try:
        if not assembly_codes or not isinstance(assembly_codes, list):
            return jsonify({'success': False, 'message': 'Assembly Code 목록이 비어있습니다'}), 400
        
        # 100개 제한 확인
        if len(assembly_codes) > 100:
            return jsonify({
                'success': False, 
                'message': f'최대 100개까지만 업로드 가능합니다. 현재: {len(assembly_codes)}개'
            }), 400
        
        logger.debug(f"Assembly Code 목록 업로드 시작: {len(assembly_codes)}개")
        
        # 데이터베이스 연결
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # 각 Assembly Code 조회 및 유효성 검사
        valid_assemblies = []
        invalid_codes = []
        
        for code in assembly_codes:
            if not code or not str(code).strip():
                continue
                
            code = str(code).strip()
            
            try:
                cursor.execute("""
                    SELECT assembly_code, company, zone, item, weight_gross,
                           fit_up_date, final_date, arup_final_date, galv_date,
                           arup_galv_date, shot_date, paint_date, arup_paint_date
                    FROM arup_ecs 
                    WHERE assembly_code = %s
                """, (code,))
                
                result = cursor.fetchone()
                if result:
                    # 조립품 데이터 구성 (실제 arup_ecs 테이블 구조에 맞게)
                    assembly_data = {
                        'assembly_code': result[0],
                        'company': result[1] or '',
                        'zone': result[2] or '',
                        'item': result[3] or '',
                        'weight_net': float(result[4]) if result[4] else 0.0,  # 호환성을 위해 weight_net으로 유지
                        'fit_up_date': result[5],
                        'final_date': result[6],
                        'arup_final_date': result[7],
                        'galv_date': result[8],
                        'arup_galv_date': result[9],
                        'shot_date': result[10],
                        'paint_date': result[11],
                        'arup_paint_date': result[12]
                    }
                    
                    # 상태 계산 후 추가
                    assembly_data = calculate_assembly_status(assembly_data)
                    valid_assemblies.append(assembly_data)
                else:
                    invalid_codes.append(code)
                    
            except Exception as e:
                logger.error(f"Assembly Code 조회 오류 ({code}): {e}")
                invalid_codes.append(code)
        
        logger.debug(f"유효한 Assembly: {len(valid_assemblies)}개, 무효한 코드: {len(invalid_codes)}개")
        
        # 유효한 Assembly가 있으면 저장리스트에 추가
        saved_count = 0
        updated_count = 0
        
        # JWT 데이터에서 user_id와 username 추출
        user_id = current_user.get('user_id')
        username = current_user.get('username')
        logger.debug(f"JWT에서 추출한 user_id: {user_id}, username: {username}")
        
        if valid_assemblies and username and user_id:
            for assembly in valid_assemblies:
                try:
                    assembly_code = assembly['assembly_code']
                    
                    # 날짜 필드들을 명시적으로 문자열로 변환 (JSON 직렬화 오류 방지)
                    assembly_safe = dict(assembly)
                    date_fields = ['fit_up_date', 'final_date', 'arup_final_date', 'galv_date', 
                                   'arup_galv_date', 'shot_date', 'paint_date', 'arup_paint_date']
                    
                    for field in date_fields:
                        if field in assembly_safe and assembly_safe[field] is not None:
                            assembly_safe[field] = str(assembly_safe[field])
                    
                    # 이미 존재하는지 확인 (user_id 사용)
                    cursor.execute("""
                        SELECT id FROM user_saved_lists 
                        WHERE user_id = %s AND assembly_code = %s
                    """, (user_id, assembly_code))
                    
                    existing = cursor.fetchone()
                    
                    if existing:
                        # 업데이트 (덮어쓰기)
                        cursor.execute("""
                            UPDATE user_saved_lists 
                            SET assembly_data = %s, updated_at = CURRENT_TIMESTAMP
                            WHERE user_id = %s AND assembly_code = %s
                        """, (json.dumps(assembly_safe), user_id, assembly_code))
                        updated_count += 1
                    else:
                        # 새로 삽입
                        cursor.execute("""
                            INSERT INTO user_saved_lists (user_id, assembly_code, assembly_data)
                            VALUES (%s, %s, %s)
                        """, (user_id, assembly_code, json.dumps(assembly_safe)))
                        saved_count += 1
                        
                except Exception as e:
                    logger.error(f"저장리스트 추가 오류 ({assembly['assembly_code']}): {e}")
        
        connection.commit()
        connection.close()
        
        # 결과 데이터 구성
        result_data = {
            'total_uploaded': len(assembly_codes),
            'valid_count': len(valid_assemblies),
            'invalid_count': len(invalid_codes),
            'saved_count': saved_count,
            'updated_count': updated_count,
            'total_in_list': saved_count + updated_count,
            'invalid_codes': invalid_codes if invalid_codes else None
        }
        
        logger.debug(f"Assembly Code 업로드 완료: {result_data}")
        
        return jsonify({
            'success': True,
            'message': f'업로드 완료! 유효: {len(valid_assemblies)}개, 무효: {len(invalid_codes)}개',
            'data': result_data
        })
        
    except Exception as e:
        logger.error(f"Assembly Code 처리 오류: {e}")
        return jsonify({'success': False, 'message': f'서버 오류: {str(e)}'}), 500